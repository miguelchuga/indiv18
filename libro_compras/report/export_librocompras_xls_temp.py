# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl import Workbook
from datetime import datetime, date, timedelta
from openpyxl.drawing.image import Image
import pathlib
import os
from openpyxl.utils import get_column_letter

import base64
from odoo import models, fields, api

class LibroComprasXLS(models.TransientModel):
    _name = 'libro.compras.xls'
    _description = "Libro compras"

    date_from = fields.Date('Start Date', default=fields.Datetime.now)
    date_to = fields.Date('End Date', default=fields.Datetime.now)
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    statew = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    state = fields.Selection([('Asentada', 'Asentada'), ('No Asentada', 'No Asentada'), ('Todas', 'Todas')], default='Asentada')
    account_id = fields.Many2one('account.account', string='Cuenta Bancaria', required=True)
    
    def generate_file(self):

        this = self.id
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        wb = Workbook()

        ws = wb.active

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        ws.title = "libro banco"

        #ESTILOS
        alig_center = Alignment(horizontal='center',vertical='center')
        alig_right = Alignment(horizontal='right', vertical='center')
        alig_left = Alignment(horizontal='left', vertical='center')
        font_black_12 = Font(name='Calibri',size=12, bold=False, italic=False, color='00000000')
        font_white_12 = Font(name='Calibri',size=12, bold=False, italic=False, color='00FFFFFF')
        font_black_12_bold = Font(name='Calibri',size=12, bold=True, italic=False, color='00000000')
        font_white_12_bold = Font(name='Calibri',size=12, bold=True, italic=False, color='00FFFFFF')
        fill_blue = PatternFill("solid", fgColor="00000080")
        fill_black = PatternFill("solid", fgColor="00000000")
        fill_white = PatternFill("solid", fgColor="00FFFFFF")
        fill_cyan = PatternFill("solid", fgColor="0000FFFF")
        thin = Side(border_style="thin", color="00000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        styleTitle = NamedStyle(name='styleTitle')
        styleTitle.alignment = alig_center
        # styleTitle.fill = fill_white
        styleTitle.font = font_black_12_bold

        styleTitleHead = NamedStyle(name='styleTitleHead')
        styleTitleHead.alignment = alig_left
        styleTitleHead.fill = fill_cyan
        styleTitleHead.font = font_black_12_bold
        styleTitleHead.border = border

        styleTitleFood = NamedStyle(name='styleTitleFood')
        styleTitleFood.alignment = alig_left
        styleTitleFood.fill = fill_white
        styleTitleFood.font = font_black_12_bold
        styleTitleFood.border = border

        styleTitleFoodNumber = NamedStyle(name='styleTitleFoodNumber')
        styleTitleFoodNumber.alignment = alig_right
        styleTitleFoodNumber.fill = fill_white
        styleTitleFoodNumber.font = font_black_12_bold
        styleTitleFoodNumber.border = border
        styleTitleFoodNumber.number_format = '#,##0.00'

        styleTitleFoodNumberInteger = NamedStyle(name='styleTitleFoodNumberInteger')
        styleTitleFoodNumberInteger.alignment = alig_right
        styleTitleFoodNumberInteger.fill = fill_white
        styleTitleFoodNumberInteger.font = font_black_12_bold
        styleTitleFoodNumberInteger.border = border

        styleCellChar = NamedStyle(name='styleCellChar')
        styleCellChar.alignment = alig_left
        styleCellChar.fill = fill_white
        styleCellChar.font = font_black_12
        styleCellChar.border = border

        styleCellNumber = NamedStyle(name='styleCellNumber')
        styleCellNumber.alignment = alig_left
        styleCellNumber.fill = fill_white
        styleCellNumber.font = font_black_12
        styleCellNumber.border = border
        styleCellNumber.number_format = '#,##0.00'


        if self.date_from:
            fs = self.date_from.strftime('%Y-%m-%d').split('-')
            _fecha_desde = ((fs[2]) + "-" + fs[1] + "-" + fs[0])

        if self.date_to:
            fs = self.date_to.strftime('%Y-%m-%d').split('-')
            _fecha_hasta = ((fs[2]) + "-" + fs[1] + "-" + fs[0])


        


        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")

        binary_data = spreadsheet_file.read()
        
        spreadsheet_file.close()
        
        out = base64.b64encode(binary_data)

        self.write({
            'statew': 'get',
            'name': "libro_banco.xlsx",
            'data': out
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'libro.banco.xls',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this,
            'views': [(False, 'form')],
            'target': 'new'
        }