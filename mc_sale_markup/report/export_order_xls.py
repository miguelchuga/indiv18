# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl import Workbook

import base64

from odoo import models, fields, api

class GenerateSaleOrderWizard(models.TransientModel):
    
    _name = "export.report.sale.order.markup"

    fecha_desde = fields.Date(string="Fecha Desde", required=True)
    fecha_hasta = fields.Date(string="Fecha Hasta", required=True)
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')

    def generate_file(self):
        
        this = self.id
        
        fileobj = NamedTemporaryFile('w+b')
        
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

        wb = Workbook()

        ws = wb.active

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        #ESTILOS
        alig_center = Alignment(horizontal='center',vertical='center')
        alig_right = Alignment(horizontal='right', vertical='center')
        alig_left = Alignment(horizontal='left', vertical='center')
        font_black_12 = Font(name='Calibri',size=12, bold=False, italic=False, color='00000000')
        font_white_12 = Font(name='Calibri',size=12, bold=False, italic=False, color='00FFFFFF')
        font_black_12_bold = Font(name='Calibri',size=12, bold=True, italic=False, color='00000000')
        font_white_12_bold = Font(name='Calibri',size=12, bold=True, italic=False, color='00FFFFFF')
        fill_blue = PatternFill("solid", fgColor="00000080")
        fill_black = PatternFill("solid", fgColor="00000000")
        fill_white = PatternFill("solid", fgColor="00FFFFFF")
        fill_cyan = PatternFill("solid", fgColor="0000FFFF")
        thin = Side(border_style="thin", color="00000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        styleTitle = NamedStyle(name='styleTitle')
        styleTitle.alignment = alig_center
        # styleTitle.fill = fill_white
        styleTitle.font = font_black_12_bold

        styleTitleHead = NamedStyle(name='styleTitleHead')
        styleTitleHead.alignment = alig_center
        styleTitleHead.fill = fill_blue
        styleTitleHead.font = font_white_12_bold
        styleTitleHead.border = border

        styleTitleFood = NamedStyle(name='styleTitleFood')
        styleTitleFood.alignment = alig_left
        styleTitleFood.fill = fill_white
        styleTitleFood.font = font_black_12_bold
        styleTitleFood.border = border

        styleTitleFoodNumber = NamedStyle(name='styleTitleFoodNumber')
        styleTitleFoodNumber.alignment = alig_right
        styleTitleFoodNumber.fill = fill_white
        styleTitleFoodNumber.font = font_black_12_bold
        styleTitleFoodNumber.border = border
        styleTitleFoodNumber.number_format = '#,##0.00'

        styleTitleFoodNumberInteger = NamedStyle(name='styleTitleFoodNumberInteger')
        styleTitleFoodNumberInteger.alignment = alig_right
        styleTitleFoodNumberInteger.fill = fill_white
        styleTitleFoodNumberInteger.font = font_black_12_bold
        styleTitleFoodNumberInteger.border = border

        styleCellChar = NamedStyle(name='styleCellChar')
        styleCellChar.alignment = alig_left
        styleCellChar.fill = fill_white
        styleCellChar.font = font_black_12
        styleCellChar.border = border

        styleCellNumber = NamedStyle(name='styleCellNumber')
        styleCellNumber.alignment = alig_right
        styleCellNumber.fill = fill_white
        styleCellNumber.font = font_black_12
        styleCellNumber.border = border
        styleCellNumber.number_format = '#,##0.00'

        styleCellNumberMarkup = NamedStyle(name='styleCellNumberMarkup')
        styleCellNumberMarkup.alignment = alig_right
        styleCellNumberMarkup.fill = fill_white
        styleCellNumberMarkup.font = font_black_12
        styleCellNumberMarkup.border = border
        styleCellNumberMarkup.number_format = '#,##0.0000'

        styleCellDate = NamedStyle(name='styleCellDate')
        styleCellDate.alignment = alig_left
        styleCellDate.fill = fill_white
        styleCellDate.font = font_black_12
        styleCellDate.border = border
        styleCellDate.number_format = 'DD-MM-YYYY'

        if self.fecha_desde:
            fs = self.fecha_desde.strftime('%Y-%m-%d').split('-')
            _fecha_desde = ((fs[2]) + "/" + fs[1] + "/" + fs[0])

        if self.fecha_hasta:
            fs = self.fecha_hasta.strftime('%Y-%m-%d').split('-')
            _fecha_hasta = ((fs[2]) + "/" + fs[1] + "/" + fs[0])

        ws.title = "Ordenes de Venta"

        ws.merge_cells('A1:H1')
        ws.merge_cells('A2:H2')
        ws.merge_cells('A3:H3')
        ws.merge_cells('A4:H4')

        row = 1
        col = 1
        # Titulo
        ws.cell(row = row,column = col).style = styleTitle
        ws.cell(row = row,column = col).value = 'ORDENES DE VENTA DETALLADAS'
        row += 1
        ws.cell(row = row,column = col).style = styleTitle
        ws.cell(row = row,column = col).value = self.env.user.company_id.name
        row += 1
        ws.cell(row = row,column = col).style = styleTitle
        ws.cell(row = row,column = col).value = 'Del: ' + _fecha_desde + ' Al: ' + _fecha_hasta
        row += 2

        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = '# ORDEN'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'FECHA'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'CLIENTE'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'PRODUCTO'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'DESCRIPCIÓN'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'MARKUP'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'COSTE'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'CANTIDAD'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'PRECIO UNITARIO'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = '# FACTURA'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'FECHA FACTURA'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'TOTAL SIN IVA'
        col += 1
        ws.cell(row = row,column = col).style = styleTitleHead
        ws.cell(row = row,column = col).value = 'TOTAL CON IVA'
        col = 1
        row += 1

        orders = self.env['sale.order'].search([('date_order','>=',self.fecha_desde),('date_order','<=',self.fecha_hasta),('state','=','sale')],order="date_order asc")


        for line in orders:
                        
            if line.invoice_count > 0:

                ws.cell(row = row,column = col).style = styleCellChar
                ws.cell(row = row,column = col).value = line.name
                col += 1
                ws.cell(row = row,column = col).style = styleCellDate
                ws.cell(row = row,column = col).value = line.date_order
                col += 1
                ws.cell(row = row,column = col).style = styleCellChar
                ws.cell(row = row,column = col).value = line.partner_id.name
                col += 1

                row_end_so = 0
                row_ini_so = row
                for li in line.order_line:
                    ws.cell(row = row,column = col).style = styleCellChar
                    ws.cell(row = row,column = col).value = li.product_id.display_name
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellChar
                    ws.cell(row = row,column = col).value = li.name
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumberMarkup
                    ws.cell(row = row,column = col).value = li.markup
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumber
                    ws.cell(row = row,column = col).value = li.standard_price
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumber
                    ws.cell(row = row,column = col).value = li.product_uom_qty
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumber
                    ws.cell(row = row,column = col).value = li.price_unit
                    row_end_so = row
                    col -= 5
                    row += 1
                
                row_end_inv = 0
                row = row_ini_so
                col += 6
                for fac in line.invoice_ids:
                    ws.cell(row = row,column = col).style = styleCellChar
                    ws.cell(row = row,column = col).value = fac.name
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellDate
                    ws.cell(row = row,column = col).value = fac.invoice_date
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumber
                    ws.cell(row = row,column = col).value = fac.amount_untaxed
                    col += 1
                    ws.cell(row = row,column = col).style = styleCellNumber
                    ws.cell(row = row,column = col).value = fac.amount_total
                    row_end_inv = row
                    col -= 3
                    row += 1

                row = row_end_so if row_end_so > row_end_inv else row_end_inv

                row += 2
                col = 1


        # Establece ancho de columnas.
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18


        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")

        binary_data = spreadsheet_file.read()
        
        spreadsheet_file.close()
        
        out = base64.b64encode(binary_data)

        self.write({
            'state': 'get',
            'name': "sale_order_details.xlsx",
            'data': out
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.report.sale.order.markup',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this,
            'views': [(False, 'form')],
            'target': 'new'
        }